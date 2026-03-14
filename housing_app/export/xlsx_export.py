from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


DEFAULT_FIELDNAMES = ["id", "url", "locality", "rooms", "area_m2", "price_total_zl", "price_per_m2_zl"]
NUMERIC_FIELDS = {"rooms", "area_m2", "price_total_zl", "price_per_m2_zl"}


def save_xlsx(listings: list[dict], path: Path, fieldnames: list[str] | None = None) -> None:
    fieldnames = fieldnames or DEFAULT_FIELDNAMES
    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"

    bold = Font(bold=True)
    for col_idx, name in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = bold

    for row_idx, listing in enumerate(listings, start=2):
        for col_idx, field in enumerate(fieldnames, start=1):
            value = listing.get(field, "")
            if field in NUMERIC_FIELDS and value != "":
                value = int(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, field in enumerate(fieldnames, start=1):
        max_len = len(field)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell_len = len(str(row[0].value)) if row[0].value is not None else 0
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    wb.save(path)
    print(f"Saved {len(listings)} listings to {path}")
